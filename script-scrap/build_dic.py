# -*- coding: utf-8 -*-
"""Gera o dicionário de dados consolidado das bases de Emendas Parlamentares
do Portal da Transparência (CGU) em um único arquivo .xlsx, uma aba por base."""

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _pasta_projeto() -> Path:
    """Acha `database_portal/` subindo a arvore, igual aos outros scripts."""
    do_ambiente = os.environ.get("EMENDAS_PROJETO")
    if do_ambiente:
        return Path(do_ambiente).expanduser()
    aqui = Path(__file__).resolve()
    for pasta in aqui.parents:
        if (pasta / "raw").is_dir() or (pasta / "processed").is_dir():
            return pasta
    return aqui.parent.parent


OUT = str(_pasta_projeto() / "metadata" / "dicionario_emendas_parlamentares.xlsx")

# Cada campo: (Campo, Descrição, Grupo, Chave)
# Grupo: Identificação | Autoria | Localidade | Classificação Orçamentária |
#        Unidade Executora | Favorecido | Documento | Convênio | Valores | Datas

D_EMENDA = [
 ("Código da Emenda","Identificador da emenda parlamentar, composto por 12 dígitos: 4 do ano da emenda + 4 do código do autor + 4 do número da emenda do autor.","Identificação","PK"),
 ("Ano da Emenda","Ano em que a emenda foi proposta.","Identificação",""),
 ("Tipo de Emenda","Descreve o tipo de emenda parlamentar.","Identificação",""),
 ("Código de Autor da Emenda","Código do autor da emenda parlamentar, conforme registrado no Sistema de Administração Financeira do Governo Federal - SIAFI.","Autoria","FK"),
 ("Nome do Autor da Emenda","Nome do autor da emenda parlamentar, conforme registrado no Sistema de Administração Financeira do Governo Federal - SIAFI.","Autoria",""),
 ("Número da Emenda","Número da emenda parlamentar, conforme registrado no Sistema de Administração Financeira do Governo Federal - SIAFI.","Identificação",""),
 ("Localidade do Gasto","Atributo do Plano de Trabalho que indica, durante a execução da despesa, a região onde a despesa ocorre.","Localidade",""),
 ("Código Município IBGE","Código IBGE do município de destinação do recurso. Este campo poderá estar em branco, a depender da localidade de aplicação.","Localidade","FK"),
 ("Município","Nome do município de destinação do recurso. Este campo poderá estar em branco, a depender da localidade de aplicação.","Localidade",""),
 ("Código UF IBGE","Código IBGE do estado de destinação do recurso. Este campo poderá estar em branco, a depender da localidade de aplicação.","Localidade","FK"),
 ("UF","Nome do estado de destinação do recurso. Este campo poderá estar sem informação, a depender da localidade de aplicação.","Localidade",""),
 ("Região","Região de destinação do recurso.","Localidade",""),
 ("Código Função","Código da Função em que foi classificada a despesa. Função – representa o maior nível de agregação das diversas áreas de atuação do setor público. Reflete a competência institucional do órgão (ex.: cultura, educação, saúde, defesa), guardando relação com os respectivos Ministérios. Fonte: Manual Técnico do Orçamento.","Classificação Orçamentária","FK"),
 ("Nome Função","Nome da Função em que foi classificada a despesa associada à emenda parlamentar.","Classificação Orçamentária",""),
 ("Código Subfunção","Código da Subfunção em que foi classificada a despesa. Subfunção – nível de agregação imediatamente inferior à função, evidencia a natureza da atuação governamental. Conforme a Portaria nº 42, de 14 de abril de 1999, é possível combinar subfunções a funções diferentes daquelas a elas diretamente relacionadas (matricialidade). Fonte: Manual Técnico do Orçamento.","Classificação Orçamentária","FK"),
 ("Nome Subfunção","Nome da subfunção em que foi classificada a despesa associada à emenda parlamentar.","Classificação Orçamentária",""),
 ("Código Programa","Código do Programa em que foi classificada a despesa. Toda ação do Governo está estruturada em programas orientados para a realização dos objetivos estratégicos definidos para o período do PPA (quatro anos). Programa Temático: expressa e orienta a entrega de bens e serviços à sociedade. Programa de Gestão, Manutenção e Serviços ao Estado: expressa e orienta as ações destinadas ao apoio, à gestão e à manutenção da atuação governamental.","Classificação Orçamentária","FK"),
 ("Nome Programa","Nome do Programa em que foi classificada a despesa. Fonte: Manual Técnico do Orçamento.","Classificação Orçamentária",""),
 ("Código Ação","Código da Ação Orçamentária em que foi classificada a despesa. Ação Orçamentária: operação da qual resultam produtos (bens ou serviços) que contribuem para atender ao objetivo de um programa. Incluem-se também no conceito de ação as transferências obrigatórias ou voluntárias a outros entes da Federação e a pessoas físicas e jurídicas, na forma de subsídios, subvenções, auxílios, contribuições, entre outros, e os financiamentos. Fonte: Manual Técnico do Orçamento.","Classificação Orçamentária","FK"),
 ("Nome Ação","Nome da ação orçamentária em que foi classificada a despesa.","Classificação Orçamentária",""),
 ("Código Plano Orçamentário","O PO é uma identificação orçamentária, de caráter gerencial (não constante da LOA), vinculada à ação orçamentária, que tem por finalidade permitir que, tanto a elaboração do orçamento quanto o acompanhamento físico e financeiro da execução, ocorram num nível mais detalhado do que o do subtítulo/localizador de gasto.","Classificação Orçamentária","FK"),
 ("Nome Plano Orçamentário","Descrição do Plano Orçamentário.","Classificação Orçamentária",""),
 ("Valor Empenhado","Valor empenhado para a emenda.","Valores",""),
 ("Valor Liquidado","Valor liquidado para a emenda.","Valores",""),
 ("Valor Pago","Valor pago para a emenda.","Valores",""),
 ("Valor Restos A Pagar Inscritos","Valor inscrito em restos a pagar para a emenda.","Valores",""),
 ("Valor Restos A Pagar Cancelados","Valor cancelado das inscrições em restos a pagar para a emenda.","Valores",""),
 ("Valor Restos A Pagar Pagos","Valor pago em restos a pagar para a emenda.","Valores",""),
]

D_CONVENIOS = [
 ("Código da Emenda","Identificador da emenda parlamentar, composto por 12 dígitos: 4 do ano da emenda + 4 do código do autor + 4 do número da emenda do autor.","Identificação","FK"),
 ("Código Função","Código da Função em que foi classificada a despesa associada à emenda parlamentar. Função – representa o maior nível de agregação das diversas áreas de atuação do setor público. Reflete a competência institucional do órgão (ex.: cultura, educação, saúde, defesa), guardando relação com os respectivos Ministérios. Fonte: Manual Técnico do Orçamento.","Classificação Orçamentária","FK"),
 ("Nome Função","Nome da Função em que foi classificada a despesa associada à emenda parlamentar.","Classificação Orçamentária",""),
 ("Código Subfunção","Código da Subfunção em que foi classificada a despesa associada à emenda parlamentar. Subfunção – nível de agregação imediatamente inferior à função, evidencia a natureza da atuação governamental. Conforme a Portaria nº 42, de 14 de abril de 1999, é possível combinar subfunções a funções diferentes daquelas a elas diretamente relacionadas (matricialidade). Fonte: Manual Técnico do Orçamento.","Classificação Orçamentária","FK"),
 ("Nome Subfunção","Nome da subfunção em que foi classificada a despesa associada à emenda parlamentar.","Classificação Orçamentária",""),
 ("Localidade do Gasto","Atributo do Plano de Trabalho que indica, durante a execução da despesa, a região onde a despesa ocorre.","Localidade",""),
 ("Tipo de Emenda","Descreve o tipo de emenda parlamentar.","Identificação",""),
 ("Data Publicação Convênio","Data de publicação do convênio.","Datas",""),
 ("Convenente","Órgão da administração direta, autárquica ou fundacional, empresa pública ou sociedade de economia mista, de qualquer esfera de governo, ou organização particular com a qual a administração federal pactua a execução de programa, projeto ou atividade, ou evento mediante a celebração de convênio. É quem recebe os recursos do Governo Federal.","Convênio",""),
 ("Objeto Convênio","Aquilo pactuado entre o Governo Federal concedente e o convenente beneficiado no município.","Convênio",""),
 ("Número Convênio","Número que identifica o convênio.","Convênio","PK"),
 ("Valor Convênio","É o valor correspondente à participação do concedente. É adicionado ao valor original do convênio a parcela (999) que corresponde a rendimento de aplicação financeira, quando for o caso.","Valores",""),
]

D_FAVORECIDO = [
 ("Código da Emenda","Identificador da emenda parlamentar, composto por 12 dígitos: 4 do ano da emenda + 4 do código do autor + 4 do número da emenda do autor.","Identificação","FK"),
 ("Código do Autor da Emenda","Código do autor da emenda parlamentar, conforme registrado no Sistema de Administração Financeira do Governo Federal - SIAFI.","Autoria","FK"),
 ("Nome do Autor da Emenda","Nome do autor da emenda parlamentar, conforme registrado no Sistema de Administração Financeira do Governo Federal - SIAFI.","Autoria",""),
 ("Número da Emenda","Número da emenda parlamentar, conforme registrado no Sistema de Administração Financeira do Governo Federal - SIAFI.","Identificação",""),
 ("Tipo de Emenda","Descreve o tipo de emenda parlamentar.","Identificação",""),
 ("Ano/Mês","Ano e mês em que foi realizado o lançamento.","Datas",""),
 ("Código do Favorecido","Código do favorecido do pagamento realizado. Favorecidos: entes governamentais, entidades sem fins lucrativos, demais pessoas jurídicas ou pessoas físicas que receberam transferências de recursos públicos federais, independentemente da origem desses valores. Fonte: Controladoria-Geral da União.","Favorecido","FK"),
 ("Favorecido","Nome do favorecido do pagamento realizado.","Favorecido",""),
 ("Natureza Jurídica","Natureza jurídica do favorecido.","Favorecido",""),
 ("Tipo Favorecido","Informa se o favorecido é Pessoa Física ou Pessoa Jurídica.","Favorecido",""),
 ("UF Favorecido","Unidade Federativa do favorecido do recurso.","Favorecido",""),
 ("Município Favorecido","Nome do município do favorecido do recurso.","Favorecido",""),
 ("Valor Recebido","Valor recebido pelo favorecido.","Valores",""),
]

D_DOCUMENTOS = [
 ("Código da Emenda","Identificador da emenda parlamentar, composto por 12 dígitos: 4 do ano da emenda + 4 do código do autor + 4 do número da emenda do autor.","Identificação","FK"),
 ("Ano da Emenda","Ano em que a emenda foi proposta.","Identificação",""),
 ("Código do Autor da Emenda","Código do autor da emenda parlamentar, conforme registrado no Sistema de Administração Financeira do Governo Federal - SIAFI.","Autoria","FK"),
 ("Nome do Autor da Emenda","Nome do autor da emenda parlamentar, conforme registrado no Sistema de Administração Financeira do Governo Federal - SIAFI.","Autoria",""),
 ("Número da Emenda","Número da emenda parlamentar, conforme registrado no Sistema de Administração Financeira do Governo Federal - SIAFI.","Identificação",""),
 ("Código Função","Código da Função em que foi classificada a despesa associada à emenda parlamentar. Função – representa o maior nível de agregação das diversas áreas de atuação do setor público. Reflete a competência institucional do órgão (ex.: cultura, educação, saúde, defesa), guardando relação com os respectivos Ministérios. Fonte: Manual Técnico do Orçamento.","Classificação Orçamentária","FK"),
 ("Valor Empenhado","Valor empenhado pelo documento para a emenda. Válido apenas para documento de empenho.","Valores",""),
 ("Valor Pago","Valor pago pelo documento para a emenda. Válido apenas para documento de pagamento.","Valores",""),
 ("Tipo de Emenda","Descreve o tipo de emenda parlamentar.","Identificação",""),
 ("Data Documento","Data do documento: para empenhos até 2020, liquidações e pagamentos, corresponde à data de emissão do documento. Para empenhos a partir de 2021, sujeitos à alteração de valor, corresponde à data da última operação realizada no documento.","Datas",""),
 ("Código Documento","Código que identifica unicamente um documento de despesa no Sistema Integrado de Administração Financeira do Governo Federal (SIAFI).","Documento","PK"),
 ("Localidade de Aplicação do Recurso","Localidade em que o recurso foi aplicado. Caso a natureza jurídica do favorecido e a modalidade de aplicação sejam uma das definidas abaixo, a localidade de aplicação do recurso será a localidade do favorecido. Caso contrário, a localidade de aplicação do recurso será a Regionalização do Gasto, atributo do Plano de Trabalho que indica, durante a execução da despesa, a região onde a despesa ocorre. Localidade do favorecido: 1. Modalidades 40 e 41 – Transferências a Municípios – Fundo a Fundo e Transferências a Municípios, cuja natureza jurídica do favorecido seja \"1244 - Município Administração Pública\" ou \"1201 - Fundo Público\". 2. Modalidade 32 – Execução Orçamentária Delegada a Estados e ao Distrito Federal, cuja natureza jurídica do favorecido seja \"1236 - Estado ou Distrito Federal\" ou \"1023 - Órgão Público do Poder Executivo Estadual ou do Distrito Federal\". 3. Modalidade 31 – Transferências a Estados e ao Distrito Federal – Fundo a Fundo, cuja natureza jurídica do favorecido seja \"1201 - Fundo Público\". 4. Modalidade 30 – Transferências a Estados e ao Distrito Federal, cuja natureza jurídica do favorecido seja \"1201 - Fundo Público\", \"1236 - Estado ou Distrito Federal\" ou \"1112 - Autarquia Estadual ou do Distrito Federal\".","Localidade",""),
 ("UF de Aplicação do Recurso","Estado em que o recurso foi aplicado, seguindo a regra de localidade do recurso.","Localidade",""),
 ("Município de Aplicação do Recurso","Município em que o recurso foi aplicado, seguindo a regra de localidade do recurso.","Localidade",""),
 ("Código IBGE do Município de Aplicação do Recurso","Código IBGE do município em que o recurso foi aplicado, seguindo a regra de localidade do recurso.","Localidade","FK"),
 ("Fase da Despesa","Fase da despesa do documento de despesa: empenho, liquidação ou pagamento.","Documento",""),
 ("Código Favorecido","Código do favorecido do documento de despesa. Favorecidos: entes governamentais, entidades sem fins lucrativos, demais pessoas jurídicas ou pessoas físicas que receberam transferências de recursos públicos federais, independentemente da origem desses valores.","Favorecido","FK"),
 ("Favorecido","Nome do destinatário do documento.","Favorecido",""),
 ("Tipo Favorecido","Informa se o favorecido é Pessoa Física, Pessoa Jurídica ou outros (exemplo: favorecido no exterior).","Favorecido",""),
 ("UF Favorecido","UF do favorecido do documento de despesa.","Favorecido",""),
 ("Município Favorecido","Município do favorecido do documento de despesa.","Favorecido",""),
 ("Código UG","Código da Unidade Orçamentária ou Administrativa que realiza atos de gestão orçamentária, financeira e/ou patrimonial.","Unidade Executora","FK"),
 ("UG","Nome da Unidade Gestora.","Unidade Executora",""),
 ("Código Unidade Orçamentária","Código da Unidade Orçamentária.","Unidade Executora","FK"),
 ("Unidade Orçamentária","Nome da Unidade Orçamentária.","Unidade Executora",""),
 ("Código Órgão SIAFI","Código SIAFI do Órgão.","Unidade Executora","FK"),
 ("Órgão","Nome do Órgão.","Unidade Executora",""),
 ("Código Órgão Superior SIAFI","Código SIAFI do Órgão Superior.","Unidade Executora","FK"),
 ("Órgão Superior","Nome do Órgão Superior.","Unidade Executora",""),
 ("Código Grupo Despesa","Código que indica em qual classe de gasto será realizada a despesa. O Grupo de Despesa é um agregador de Elemento de Despesa com as mesmas características quanto ao objeto de gasto. Pode assumir os valores \"1\" (Pessoal e Encargos Sociais), \"2\" (Juros e Encargos da Dívida), \"3\" (Outras Despesas Correntes), \"4\" (Investimentos), \"5\" (Inversões Financeiras) ou \"6\" (Amortização da Dívida). Fonte: Manual Técnico do Orçamento (SOF).","Classificação Orçamentária","FK"),
 ("Grupo Despesa","Descrição do Grupo de Despesa. Pode assumir os valores \"Pessoal e Encargos Sociais\", \"Juros e Encargos da Dívida\", \"Outras Despesas Correntes\", \"Investimentos\", \"Inversões Financeiras\" ou \"Amortização da Dívida\".","Classificação Orçamentária",""),
 ("Código Elemento Despesa","Código que tem por finalidade identificar os objetos de gasto, tais como vencimentos e vantagens fixas, juros, diárias, material de consumo, serviços de terceiros prestados sob qualquer forma, subvenções sociais, obras e instalações, equipamentos e material permanente, auxílios, amortização e outros que a Administração Pública utiliza para a consecução de seus fins. Fonte: Manual Técnico do Orçamento (SOF).","Classificação Orçamentária","FK"),
 ("Elemento Despesa","Descrição do Elemento de Despesa.","Classificação Orçamentária",""),
 ("Código Modalidade Aplicação Despesa","Código que indica de que forma os recursos serão aplicados: diretamente, pela unidade detentora do crédito orçamentário; indiretamente mediante transferência, por outras esferas de governo, seus órgãos, fundos ou entidades, ou por entidades privadas; ou indiretamente mediante delegação, por outros entes federativos ou consórcios públicos. Compõe o campo natureza da despesa (são os 3º e 4º dígitos deste campo do código) e possibilita a eliminação da dupla contagem dos recursos transferidos ou descentralizados.","Classificação Orçamentária","FK"),
 ("Modalidade Aplicação Despesa","Descrição da Modalidade de Despesa que indica de que forma os recursos serão aplicados.","Classificação Orçamentária",""),
 ("Código Plano Orçamentário","Código de identificação orçamentária, de caráter gerencial (não constante da LOA), vinculada à ação orçamentária, que tem por finalidade permitir que, tanto a elaboração do orçamento quanto o acompanhamento físico e financeiro da execução, ocorram num nível mais detalhado do que o do subtítulo/localizador de gasto. Fonte: Manual Técnico do Orçamento (SOF).","Classificação Orçamentária","FK"),
 ("Plano Orçamentário","Descrição do Plano Orçamentário.","Classificação Orçamentária",""),
 ("Função","Nome da Função em que foi classificada a despesa.","Classificação Orçamentária",""),
 ("Código Subfunção","Código da Subfunção em que foi classificada a despesa associada à emenda parlamentar. Subfunção – nível de agregação imediatamente inferior à função, evidencia a natureza da atuação governamental. Conforme a Portaria nº 42, de 14 de abril de 1999, é possível combinar subfunções a funções diferentes daquelas a elas diretamente relacionadas (matricialidade). Fonte: Manual Técnico do Orçamento.","Classificação Orçamentária","FK"),
 ("Subfunção","Nome da subfunção em que foi classificada a despesa.","Classificação Orçamentária",""),
 ("Código Programa","Código do Programa em que foi classificada a despesa. Toda ação do Governo está estruturada em programas orientados para a realização dos objetivos estratégicos definidos para o período do PPA (quatro anos). Programa Temático: expressa e orienta a entrega de bens e serviços à sociedade. Programa de Gestão, Manutenção e Serviços ao Estado: expressa e orienta as ações destinadas ao apoio, à gestão e à manutenção da atuação governamental.","Classificação Orçamentária","FK"),
 ("Programa","Nome do Programa em que foi classificada a despesa. Fonte: Manual Técnico do Orçamento.","Classificação Orçamentária",""),
 ("Código Ação","Código da Ação Orçamentária em que foi classificada a despesa. Ação Orçamentária: operação da qual resultam produtos (bens ou serviços) que contribuem para atender ao objetivo de um programa. Incluem-se também no conceito de ação as transferências obrigatórias ou voluntárias a outros entes da Federação e a pessoas físicas e jurídicas, na forma de subsídios, subvenções, auxílios, contribuições, entre outros, e os financiamentos. Fonte: Manual Técnico do Orçamento.","Classificação Orçamentária","FK"),
 ("Ação","Nome da ação orçamentária em que foi classificada a despesa.","Classificação Orçamentária",""),
 ("Linguagem Cidadã","Nomes mais intuitivos pelos quais as ações governamentais são apresentadas aos cidadãos. Exemplo: Transferência de Renda Diretamente às Famílias em Condição de Pobreza e Extrema Pobreza (Bolsa Família).","Classificação Orçamentária",""),
 ("Código Subtítulo (Localizador)","Código utilizado especialmente para especificar a localização física integral ou parcial das ações orçamentárias.","Classificação Orçamentária","FK"),
 ("Subtítulo (Localizador)","Descrição do Subtítulo.","Classificação Orçamentária",""),
 ("Possui convênio?","Informa se há algum convênio relacionado ao documento de despesa.","Convênio",""),
]

D_APOIAMENTOS = [
 ("Código Apoiador","Código do parlamentar apoiador/solicitante do empenho.","Autoria","FK"),
 ("Apoiador","Parlamentar apoiador/solicitante do empenho.","Autoria",""),
 ("Data do Apoio","Data de registro do apoio no sistema. ATENÇÃO: no arquivo de 2022 todas as datas caem entre 2025 e 2026, em formato ISO com hora (2025-06-13 12:18:16) — parece ser o carimbo de carga, não a data do ato político. Para ordenar no tempo, use Data última movimentação Empenho.","Datas",""),
 ("Data Retirada do Apoio","Data de registro de retirada do apoio.","Datas",""),
 ("Empenho","Código que identifica unicamente um empenho no SIAFI. O dicionário oficial descreve 6 dígitos de UG + 5 de Gestão + 11 de Empenho = 22 caracteres, mas o arquivo real traz 23: UG (6) + Gestão (5) + ano (4) + tipo (2, sempre 'NE') + número (6). Ex.: 110594000012022NE000135. É idêntico ao Código Documento da base Por Documento de Despesa (fase empenho) — a junção entre as duas é direta.","Documento","PK"),
 ("Data última movimentação Empenho","Data da última movimentação do empenho. Formato DD/MM/AAAA; pode vir como 'Sem informação'.","Datas",""),
 ("Código favorecido","Código do destinatário do documento, por exemplo, seu CNPJ. Os favorecidos podem ser entes governamentais, entidades sem fins lucrativos ou demais pessoas jurídicas e pessoas físicas que recebem recursos públicos federais, independentemente da origem desses valores.","Favorecido","FK"),
 ("Favorecido","Nome do favorecido do empenho realizado.","Favorecido",""),
 ("Tipo Favorecido","Informa se o favorecido é Pessoa Física ou Pessoa Jurídica.","Favorecido",""),
 ("UF Favorecido","Unidade Federativa do favorecido do empenho.","Favorecido",""),
 ("Municipio Favorecido","Nome do município do favorecido do empenho.","Favorecido",""),
 ("Código da Emenda","Identificador da emenda parlamentar, composto por 12 dígitos: 4 do ano da emenda + 4 do código do autor + 4 do número da emenda do autor.","Identificação","FK"),
 ("Tipo de Emenda","Descreve o tipo de emenda parlamentar.","Identificação",""),
 ("Ano da Emenda","Ano em que a emenda foi proposta.","Identificação",""),
 # As quatro colunas a seguir EXISTEM no arquivo e NAO constam do dicionario
 # oficial da CGU (que lista 27 campos; o CSV real tem 31). Verificado em
 # 18/08/2026 no arquivo de 2022. Sao o que permite comparar autor formal x
 # apoiador real sem precisar juntar com a base de emendas.
 ("Código do Autor da Emenda","Código do autor da emenda conforme SIAFI. Em emendas de bancada, comissão ou relator o autor é a entidade coletiva — o parlamentar individual aparece na coluna Apoiador. AUSENTE do dicionário oficial.","Autoria","FK"),
 ("Nome do Autor da Emenda","Nome do autor da emenda conforme SIAFI. AUSENTE do dicionário oficial.","Autoria",""),
 ("Número da emenda","Número da emenda conforme SIAFI. AUSENTE do dicionário oficial.","Identificação",""),
 ("Localidade de aplicação do recurso","Localidade de destino do recurso. AUSENTE do dicionário oficial.","Localidade",""),
 ("Código UG","Código da Unidade Orçamentária ou Administrativa que realiza atos de gestão orçamentária, financeira e/ou patrimonial.","Unidade Executora","FK"),
 ("UG","Nome da Unidade Gestora.","Unidade Executora",""),
 ("Código Unidade Orçamentária","Código da unidade responsável pela coordenação do processo de elaboração da proposta orçamentária no seu âmbito de atuação, integrando e articulando o trabalho das suas unidades administrativas, tendo em vista a consistência da programação de sua unidade. Fonte: Manual Técnico do Orçamento (SOF).","Unidade Executora","FK"),
 ("Unidade Orçamentária","Nome da Unidade Orçamentária.","Unidade Executora",""),
 ("Código Órgão SIAFI","Código SIAFI do Órgão emitente do documento, subordinado ao Órgão Superior. Órgão Subordinado é a entidade supervisionada por um Órgão da Administração Direta.","Unidade Executora","FK"),
 ("Órgão","Nome do Órgão emitente do documento.","Unidade Executora",""),
 ("Código Órgão Superior SIAFI","Código SIAFI do Órgão Superior emitente do documento. Órgão Superior é a Unidade da Administração Direta que tem entidades por ele supervisionadas. Fonte: Manual do SIAFI.","Unidade Executora","FK"),
 ("Órgão Superior","Nome do Órgão Superior emitente do documento.","Unidade Executora",""),
 ("Código Ação","Código da Ação Orçamentária em que foi classificada a despesa. Ação Orçamentária: operação da qual resultam produtos (bens ou serviços) que contribuem para atender ao objetivo de um programa. Incluem-se também no conceito de ação as transferências obrigatórias ou voluntárias a outros entes da Federação e a pessoas físicas e jurídicas, na forma de subsídios, subvenções, auxílios, contribuições, entre outros, e os financiamentos. Fonte: Manual Técnico do Orçamento.","Classificação Orçamentária","FK"),
 ("Ação","Nome da ação orçamentária em que foi classificada a despesa.","Classificação Orçamentária",""),
 ("Valor Empenhado","Valor total do empenho (não inclui os valores cancelados).","Valores",""),
 ("Valor Cancelado","Valor total cancelado do empenho.","Valores",""),
 ("Valor Pago","Valor total pago referente ao empenho.","Valores",""),
]

SHEETS = [
 ("1. Por Emenda", "Emendas Parlamentares – Por Emenda Parlamentar",
  "Visão agregada da execução orçamentária de cada emenda, já classificada por função/subfunção/programa/ação e por localidade de destino. Uma linha por combinação emenda × classificação orçamentária × localidade.", D_EMENDA),
 ("2. Por Documento de Despesa", "Emendas Parlamentares – Por Documentos de Despesa",
  "Nível mais granular: cada documento SIAFI (empenho, liquidação ou pagamento) vinculado a uma emenda, com favorecido, unidade executora e classificação completa da despesa.", D_DOCUMENTOS),
 ("3. Por Favorecido", "Emendas Parlamentares – Por Favorecido",
  "Pagamentos agregados por favorecido e por Ano/Mês. Responde 'quem recebeu quanto' de cada emenda.", D_FAVORECIDO),
 ("4. Convênios", "Emendas Parlamentares – Convênios",
  "Convênios celebrados a partir de emendas, com convenente, objeto, número e valor do convênio.", D_CONVENIOS),
 ("5. Apoiamentos", "Apoiamentos de Emendas Parlamentares",
  "Registro de parlamentares que apoiaram/solicitaram um empenho específico (típico de emendas de bancada/relator). Uma linha por apoiador × empenho.", D_APOIAMENTOS),
]

# ---------------- estilos ----------------
FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, size=14, bold=True, color="1F3864")
SUB_FONT = Font(name=FONT, size=9, italic=True, color="404040")
BODY = Font(name=FONT, size=9)
BODY_B = Font(name=FONT, size=9, bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT_FILL = PatternFill("solid", fgColor="F2F5FA")
KEY_FILL = PatternFill("solid", fgColor="FFF2CC")

GRUPO_CORES = {
 "Identificação":"D9E1F2","Autoria":"E2EFDA","Localidade":"FCE4D6",
 "Classificação Orçamentária":"FFF2CC","Unidade Executora":"DDEBF7",
 "Favorecido":"E4DFEC","Documento":"F8CBAD","Convênio":"D9D9D9",
 "Valores":"C6E0B4","Datas":"FFE699",
}

wb = Workbook()

# ---------------- aba índice ----------------
ws = wb.active
ws.title = "Índice"
ws["A1"] = "Dicionário de Dados — Emendas Parlamentares"
ws["A1"].font = Font(name=FONT, size=16, bold=True, color="1F3864")
ws["A2"] = "Portal da Transparência · Controladoria-Geral da União (CGU) · Dicionários consultados em 06/08/2026"
ws["A2"].font = SUB_FONT

hdr = ["Aba", "Base de dados", "Granularidade / conteúdo", "Nº de variáveis", "Chave principal"]
for j, h in enumerate(hdr, 1):
    c = ws.cell(row=4, column=j, value=h)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

chaves = {
 "1. Por Emenda":"Código da Emenda (+ classificação + localidade)",
 "2. Por Documento de Despesa":"Código Documento",
 "3. Por Favorecido":"Código da Emenda + Código do Favorecido + Ano/Mês",
 "4. Convênios":"Número Convênio",
 "5. Apoiamentos":"Empenho + Código Apoiador",
}
r = 5
for tab, nome, desc, campos in SHEETS:
    vals = [tab, nome, desc, len(campos), chaves[tab]]
    for j, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = BODY_B if j == 1 else BODY
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
    if r % 2 == 1:
        for j in range(1, 6):
            ws.cell(row=r, column=j).fill = ALT_FILL
    r += 1

ws.cell(row=r+1, column=1, value="Legenda de chaves").font = BODY_B
ws.cell(row=r+2, column=1, value="PK").font = BODY_B
ws.cell(row=r+2, column=2, value="Identificador (ou parte do identificador) do registro na base.").font = BODY
ws.cell(row=r+3, column=1, value="FK").font = BODY_B
ws.cell(row=r+3, column=2, value="Campo de ligação com outra base ou com uma tabela de domínio (município IBGE, órgão SIAFI, função/programa/ação etc.).").font = BODY
ws.cell(row=r+5, column=1, value="Fonte: https://portaldatransparencia.gov.br — Dicionários de Dados de Emendas Parlamentares, Emendas Parlamentares por Documentos de Despesa e Apoiamentos de Emendas Parlamentares (versão do portal 6.4.9).").font = SUB_FONT

for col, w in zip("ABCDE", [26, 46, 60, 14, 42]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ---------------- abas de dicionário ----------------
for tab, nome, desc, campos in SHEETS:
    s = wb.create_sheet(tab)
    s["A1"] = nome
    s["A1"].font = TITLE_FONT
    s["A2"] = desc
    s["A2"].font = SUB_FONT
    s["A3"] = f"{len(campos)} variáveis"
    s["A3"].font = SUB_FONT

    heads = ["#", "Campo", "Descrição", "Grupo de variáveis", "Chave"]
    for j, h in enumerate(heads, 1):
        c = s.cell(row=5, column=j, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, (campo, descr, grupo, chave) in enumerate(campos, 1):
        row = 5 + i
        vals = [i, campo, descr, grupo, chave]
        for j, v in enumerate(vals, 1):
            c = s.cell(row=row, column=j, value=v)
            c.font = BODY_B if j == 2 else BODY
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=(j == 3),
                                    horizontal="center" if j in (1, 5) else "left")
        s.cell(row=row, column=4).fill = PatternFill("solid", fgColor=GRUPO_CORES.get(grupo, "FFFFFF"))
        if chave:
            s.cell(row=row, column=5).fill = KEY_FILL
            s.cell(row=row, column=5).font = BODY_B

    for col, w in zip("ABCDE", [5, 42, 105, 28, 8]):
        s.column_dimensions[col].width = w
    s.freeze_panes = "A6"
    s.auto_filter.ref = f"A5:E{5+len(campos)}"

wb.save(OUT)
print("OK:", OUT)
for tab, nome, desc, campos in SHEETS:
    print(f"{tab}: {len(campos)} campos")
